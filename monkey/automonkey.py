import subprocess
import time
import datetime
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import random

appstart='adb shell am start -W '  # 启动的activity
adbkillapp='adb shell am force-stop com.and.who'
cmd_activity="adb shell dumpsys activity activities | sed -En -e '/Running activities/,/Run #0/p'"
cmd_brand='adb shell "getprop | grep ro.product.brand"'  # 获取手机型号
cmd_findMonkey='adb shell ps | grep monkey'



def adb(cmds):
	amstart=subprocess.Popen(cmds,shell=True,stdout=subprocess.PIPE)
	return amstart

def Cmd_brand():
	global brands_cmd,new_s
	cmd_brands=adb(cmd_brand).stdout.read().decode('utf-8').split()[1]
	if cmd_brands is not None:
		new_s=cmd_brands.replace('[','').replace(']','')
	else:
		print('budui')
	return new_s

def killphone(pid_name):
	kill_s='adb shell kill %s'%(pid_name)
	adb(kill_s)

def action_time():
	now=time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
	return now

def creat_excel():
	file_name='App_monkey.xlsx'
	work=Workbook()
	work.save(file_name)
	work_book=load_workbook(file_name)
	sheets=work_book.active
	sheets.cell(1,1,'设备名')
	sheets.cell(1,2,'应用名')
	sheets.cell(1,3,'运行时间')
	sheets.cell(1,4,'执行结果')
	sheets.cell(1,5,'执行轮次')
	sheets.cell(1,6,'错误内容')
	return sheets,work_book

def start_monkey(pack_name):
	cmd_monkey = 'adb shell monkey -p %s -s 233 --pct-touch 70 --pct-motion 30 --ignore-crashes --ignore-timeouts ' \
	             '--monitor-native-crashes --throttle 200 -v -v 500'%(pack_name)
	monkey_s=subprocess.Popen(cmd_monkey,shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
	return monkey_s

def getpid(cmd_pids):
	pidsum = adb(cmd_pids).stdout.read().decode('utf-8').split()
	pids=[]
	if pidsum!=[]:
		for i in range(len(pidsum)):
			pids.append(pidsum[i+1])
			return pids
	else:
		print(len(pidsum))
		return pidsum
def activity_top():
	activity_s=adb(cmd_activity).stdout.read().decode().split()[7][2:]
	return activity_s

def seve_excel(sheets):
	for j in sheets[0].rows:  # we.rows 获取每一行数据
		for n in j:
			print(n.value,end="\t")  # n.value 获取单元格的值
		print()
	# 保存，save（必须要写文件名（绝对地址）默认 py 同级目录下，只支持 xlsx 格式）
	save_name='logcat/monkey_time_{0}.xlsx'.format(action_time())
	sheets[1].save(save_name)

brank=False
def stat_monke(count,pack_Name):
	dhl=input('是否禁用状态栏，输入1禁用，输入其他不禁用')
	if dhl ==1:
		adb('adb shell settings put global policy_control immersive.full=*')
	else:
		pass
	excel_s=creat_excel()
	cmd_clear='adb logcat -c'
	clear_logcat=adb(cmd_clear)
	for i in range(1,count+1):
		global brank
		monkey_sta=start_monkey(pack_Name)
		cmd_logcat='adb logcat -v time'
		file_name='logcat/'+'logcat_'+Cmd_brand()+'_'+str(i)+'_'+action_time()+'.txt'
		sub=subprocess.Popen(cmd_logcat,shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
		
		excel_s[0].cell(i+1,1,Cmd_brand())
		excel_s[0].cell(i+1,2,pack_Name)
		excel_s[0].cell(i+1,3,datetime.datetime.now().strftime("%H:%M:%S"))
		excel_s[0].cell(i+1,5,i)
		with open('%s'%(file_name),'wb') as report:
			for line in sub.stdout:
				monkey_pid=getpid(cmd_findMonkey)
				report.write(line)
				now_activity=activity_top()
				if 'com.wodi.who' == now_activity:
					print(('当前所在应用是%s，在玩吧中')%(now_activity))
				else:
					# 'com.wodi.who' not in now_activity:
					start_time=adb(appstart)
					print(str(random.randint(0,900))+'当前在所在应用是%s'%(now_activity)+'跳出应用了，已拉回')
				if 'Fatal'or 'Crash' or 'AndroidRuntime' or 'Exception'or 'Error' in str(line):
					print(str(line))
					excel_s[0].cell(i+1,6,str(line)+'\n')
					excel_s[0].cell(i+1,4,'FAIL')
				
					# monkey_pid=getpid(cmd_findMonkey)
					# for i in monkey_pid:
					# 	killphone(i)
					# break
				if len(monkey_pid)==0:
					excel_s[0].cell(i+1,4,'PASS')
					break
				else:
		
					excel_s[0].cell(i+1,4,'不知道为啥')
		excel_s[0].cell(i+1,6,file_name)
	seve_excel(excel_s)
		# if monkey_pid ==[]:
		# 	excel_s[0].cell(i+1,5,'PASS')
		# 	break
		# else:
		# 	print('运行中')
pack='com.wodi.who'
stat_monke(2,pack)
	